from openpyxl import load_workbook
from datetime import datetime
import pytz

class Event:
    def __init__(self, game_id, event_id, group, title, short_description, long_description, event_type, game_system,
                 rules_edition, min_players, max_players, age_required, experience_required, materials_required,
                 materials_details, start_datetime, duration, end_datetime, gm_names, website, email, tournament,
                 round_number, total_rounds, min_play_time, attendee_registration, cost, location, room_name,
                 table_number, special_category, tickets_available, last_modified):
        self.game_id = game_id
        self.event_id = event_id
        self.group = group
        self.title = title
        self.short_description = short_description
        self.long_description = long_description
        self.event_type = event_type
        self.game_system = game_system
        self.rules_edition = rules_edition
        self.min_players = min_players
        self.max_players = max_players
        self.age_required = age_required
        self.experience_required = experience_required
        self.materials_required = materials_required
        self.materials_details = materials_details
        self.start_datetime = self.format_datetime(start_datetime)
        self.duration = self.format_duration(duration)
        self.end_datetime = end_datetime
        self.gm_names = gm_names
        self.website = website
        self.email = email
        self.tournament = tournament
        self.round_number = round_number
        self.total_rounds = total_rounds
        self.min_play_time = min_play_time
        self.attendee_registration = attendee_registration
        self.cost = f"${float(cost):.2f}" if cost else "$0.00"
        self.location = location
        self.room_name = room_name
        self.table_number = table_number
        self.special_category = special_category
        self.tickets_available = tickets_available
        self.last_modified = last_modified

    @staticmethod
    def format_datetime(datetime_str):
        try:
            # Parse the datetime string in the format "MM/DD/YYYY HH:MM AM" as UTC
            dt = datetime.strptime(datetime_str, "%m/%d/%Y %I:%M %p")
            utc_tz = pytz.utc
            dt = utc_tz.localize(dt)  # Localize to UTC
            # Convert to Eastern Time
            local_tz = pytz.timezone("US/Eastern")
            dt = dt.astimezone(local_tz)
            # Format as "Day, HH:MM AM/PM TZ"
            return dt.strftime("%A, %I:%M %p %Z")
        except Exception as e:
            return "Invalid datetime"

    @staticmethod
    def format_duration(duration):
        try:
            hours = float(duration)
            return f"{hours:.1f} hr" if hours % 1 != 0 else f"{int(hours)} hr"
        except ValueError:
            return "Invalid duration"

    @staticmethod
    def load_events(file_path):
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]  # Extract headers from the first row
        events_dict = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Iterate over rows starting from the second
            row_data = dict(zip(headers, row))
            event = Event(
                game_id=row_data['Game ID'],
                event_id=row_data['Game ID'][-6:],  # Extract last 6 digits for event ID
                group=row_data['Group'],
                title=row_data['Title'],
                short_description=row_data['Short Description'],
                long_description=row_data['Long Description'],
                event_type=row_data['Event Type'],
                game_system=row_data['Game System'],
                rules_edition=row_data['Rules Edition'],
                min_players=row_data['Minimum Players'],
                max_players=row_data['Maximum Players'],
                age_required=row_data['Age Required'],
                experience_required=row_data['Experience Required'],
                materials_required=row_data['Materials Required'],
                materials_details=row_data['Materials Required Details'],
                start_datetime=row_data['Start Date & Time'],
                duration=row_data['Duration'],
                end_datetime=row_data['End Date & Time'],
                gm_names=row_data['GM Names'],
                website=row_data['Website'],
                email=row_data['Email'],
                tournament=row_data['Tournament?'],
                round_number=row_data['Round Number'],
                total_rounds=row_data['Total Rounds'],
                min_play_time=row_data['Minimum Play Time'],
                attendee_registration=row_data['Attendee Registration?'],
                cost=row_data['Cost $'],
                location=row_data['Location'],
                room_name=row_data['Room Name'],
                table_number=row_data['Table Number'],
                special_category=row_data['Special Category'],
                tickets_available=row_data['Tickets Available'],
                last_modified=row_data['Last Modified']
            )
            # Game ID is a full form like ENT25ND274324 - we want to use the last 6 digits to match the URL
            # e.g. https://www.gencon.com/events/274324
            events_dict[event.event_id] = event
        return events_dict
